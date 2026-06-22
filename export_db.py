import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from database.operations import MaterialDatabase
from database.models import TestCategory, TestFieldDefinition, MaterialSpec, TestValue, FieldMapping

def export_all_data(db_path='material_spec.db', output_file='material_spec_export.xlsx'):
    db = MaterialDatabase(db_path)
    session = db.get_session()
    
    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    try:
        categories = session.query(TestCategory).order_by(TestCategory.id).all()
        field_defs = session.query(TestFieldDefinition).order_by(TestFieldDefinition.category_id, TestFieldDefinition.sort_order).all()
        specs = session.query(MaterialSpec).order_by(MaterialSpec.spec_number, MaterialSpec.alloy_grade).all()
        test_values = session.query(TestValue).all()
        field_mappings = session.query(FieldMapping).all()

        cat_df = pd.DataFrame([{'ID': c.id, '名称': c.name, '代码': c.code, '描述': c.description or '', '是否启用': '是' if c.is_active else '否', '排序': c.sort_order} for c in categories])
        ws_cat = wb.active
        ws_cat.title = "测试类别"
        _write_dataframe(ws_cat, cat_df, header_fill, header_font_white, thin_border)

        fd_df = pd.DataFrame([{
            'ID': f.id,
            '类别ID': f.category_id,
            '字段名称': f.field_name,
            '字段代码': f.field_code,
            '字段类型': f.field_type,
            '默认单位': f.unit or '',
            '描述': f.description or '',
            '是否必填': '是' if f.is_required else '否',
            '排序': f.sort_order
        } for f in field_defs])
        ws_fd = wb.create_sheet("字段定义")
        _write_dataframe(ws_fd, fd_df, header_fill, header_font_white, thin_border)

        fm_df = pd.DataFrame([{
            'ID': m.id,
            '源字段名': m.source_field,
            '目标字段代码': m.target_field_code,
            '类别代码': m.category_code or '',
            '是否启用': '是' if m.is_active else '否'
        } for m in field_mappings])
        ws_fm = wb.create_sheet("字段映射")
        _write_dataframe(ws_fm, fm_df, header_fill, header_font_white, thin_border)

        value_map = {}
        for tv in test_values:
            if tv.spec_id not in value_map:
                value_map[tv.spec_id] = {}
            field_code = tv.field_definition.field_code if tv.field_definition else f'field_{tv.field_definition_id}'
            val_str = _format_test_value(tv)
            value_map[tv.spec_id][field_code] = val_str

        all_field_codes = set()
        for vals in value_map.values():
            all_field_codes.update(vals.keys())
        sorted_fields = sorted(all_field_codes)

        rows = []
        for spec in specs:
            row = {
                'ID': spec.id,
                '标准号': spec.material_spec_number,
                '类别代码': spec.test_category.code if spec.test_category else '',
                '类别名称': spec.test_category.name if spec.test_category else '',
                '牌号': spec.alloy_grade,
                '状态': spec.status,
                '规格': spec.specification,
                '取样方向': spec.sampling_direction,
                '附加条件': spec.additional_conditions or '',
                '备注': spec.remarks or ''
            }
            spec_values = value_map.get(spec.id, {})
            for fc in sorted_fields:
                row[fc] = spec_values.get(fc, '')
            rows.append(row)

        cols_order = ['ID', '标准号', '类别代码', '类别名称', '牌号', '状态', '规格', '取样方向', '附加条件', '备注'] + sorted_fields
        spec_df = pd.DataFrame(rows, columns=cols_order)
        ws_spec = wb.create_sheet("材料规格数据")
        _write_dataframe(ws_spec, spec_df, header_fill, header_font_white, thin_border)

        for col_idx in range(1, len(cols_order) + 1):
            col_letter = ws_spec.cell(row=1, column=col_idx).column_letter
            ws_spec.column_dimensions[col_letter].width = 15
        ws_spec.column_dimensions['A'].width = 8
        ws_spec.column_dimensions['B'].width = 25
        ws_spec.column_dimensions['E'].width = 20
        ws_spec.column_dimensions['J'].width = 30
        ws_spec.column_dimensions['K'].width = 30

        for ws in [ws_cat, ws_fd, ws_fm]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        wb.save(output_file)
        print(f"导出成功: {output_file}")
        print(f"- 测试类别: {len(categories)} 条")
        print(f"- 字段定义: {len(field_defs)} 条")
        print(f"- 字段映射: {len(field_mappings)} 条")
        print(f"- 材料规格: {len(specs)} 条 (含 {len(test_values)} 个测试值)")
        
        return output_file
        
    finally:
        session.close()

def _format_test_value(tv):
    if tv.string_value and tv.string_value.strip():
        return tv.string_value
    
    parts = []
    if tv.comparison:
        parts.append(tv.comparison)
    if tv.min_value:
        parts.append(tv.min_value)
    if tv.max_value:
        if tv.min_value:
            parts.append('~')
        parts.append(tv.max_value)
    if tv.number_value and not tv.min_value and not tv.max_value:
        parts.append(tv.number_value)
    
    result = ''.join(parts)
    if tv.unit:
        result += f" {tv.unit}"
    return result if result else ''

def _write_dataframe(ws, df, fill, font, border):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if r_idx == 1:
                cell.fill = fill
                cell.font = font

if __name__ == '__main__':
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'material_specs.db')
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'material_spec_export.xlsx')
    export_all_data(db_path, output_path)
